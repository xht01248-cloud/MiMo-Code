#!/usr/bin/env python3
"""Bake cached formula values into an xlsx via headless LibreOffice.

openpyxl only stores formulas as strings — it does not evaluate them. This
script hands the workbook to LibreOffice, asks it to re-serialize it as
xlsx (which triggers recalculation on the way out), then scans the result
for Excel error tokens.

Usage:
    python bake.py workbook.xlsx [--timeout 30]

Prints a JSON report to stdout.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))
from runtime.libreoffice import (  # noqa: E402
    LibreOfficeNotFound,
    invoke,
    locate_libreoffice,
)

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"ok": False, "reason": "openpyxl is not installed"}),
          file=sys.stderr)
    sys.exit(2)


ERROR_TOKENS: tuple[str, ...] = (
    "#VALUE!", "#DIV/0!", "#REF!", "#NAME?",
    "#NULL!", "#NUM!", "#N/A", "#GETTING_DATA",
)


# --- bake ------------------------------------------------------------------


class BakeFailed(RuntimeError):
    """LibreOffice could not open, recalculate, or re-emit the workbook."""


def _force_full_calc(source: Path, dest: Path) -> None:
    """Copy source → dest with `calcPr fullCalcOnLoad="1"` patched into
    xl/workbook.xml, so LibreOffice recalculates even formulas that already
    carry a stale cached value (its default for Excel files is "never
    recalculate on load")."""
    with zipfile.ZipFile(source) as zin, \
            zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item)
            if item.filename == "xl/workbook.xml":
                text = data.decode("utf-8")
                if "<calcPr" in text:
                    if "fullCalcOnLoad" in text:
                        text = re.sub(r'fullCalcOnLoad="[^"]*"',
                                      'fullCalcOnLoad="1"', text, count=1)
                    else:
                        text = text.replace(
                            "<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
                else:
                    # calcPr must follow definedNames/externalReferences/sheets
                    # in the CT_Workbook sequence; insert after the last one present.
                    for anchor in ("</definedNames>", "</externalReferences>",
                                   "</sheets>"):
                        if anchor in text:
                            text = text.replace(
                                anchor, anchor + '<calcPr fullCalcOnLoad="1"/>', 1)
                            break
                data = text.encode("utf-8")
            zout.writestr(item, data)


def bake_workbook(path: Path, timeout_seconds: float) -> None:
    """Overwrite ``path`` with a copy re-serialized by LibreOffice."""
    with tempfile.TemporaryDirectory(prefix="xlsx-bake-") as scratch:
        in_dir = Path(scratch) / "in"
        out_dir = Path(scratch) / "out"
        in_dir.mkdir()
        out_dir.mkdir()

        patched = in_dir / path.name
        _force_full_calc(path, patched)

        result = invoke(
            [
                "--headless",
                "--calc",
                "--convert-to", "xlsx",
                "--outdir", str(out_dir),
                str(patched),
            ],
            timeout=timeout_seconds,
        )
        if not result.ok:
            raise BakeFailed(
                f"soffice returned {result.returncode}: "
                f"{result.stderr.strip() or '(no stderr)'}"
            )

        produced = out_dir / path.name
        if not produced.exists():
            fallback = next(out_dir.glob("*.xlsx"), None)
            if fallback is None:
                raise BakeFailed(
                    "LibreOffice produced no xlsx. stderr: "
                    + (result.stderr.strip() or "(empty)")
                )
            produced = fallback

        shutil.move(str(produced), str(path))


# --- error collection ------------------------------------------------------


@dataclass
class ErrorBreakdown:
    token: str
    count: int
    example_locations: list[str] = field(default_factory=list)


def collect_error_cells(path: Path) -> tuple[list[ErrorBreakdown], int]:
    buckets: dict[str, list[str]] = {tok: [] for tok in ERROR_TOKENS}
    total = 0
    wb = load_workbook(path, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str) or not value.startswith("#"):
                        continue
                    for token in ERROR_TOKENS:
                        if value == token or value.startswith(token):
                            buckets[token].append(f"{sheet_name}!{cell.coordinate}")
                            total += 1
                            break
    finally:
        wb.close()

    breakdowns = [
        ErrorBreakdown(token=tok, count=len(locs), example_locations=locs[:20])
        for tok, locs in buckets.items()
        if locs
    ]
    return breakdowns, total


def count_formulas(path: Path) -> int:
    wb = load_workbook(path, data_only=False)
    try:
        total = 0
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        total += 1
        return total
    finally:
        wb.close()


# --- CLI -------------------------------------------------------------------


def _parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Recalculate and re-save an xlsx via headless LibreOffice.",
    )
    p.add_argument("workbook", type=Path)
    p.add_argument("--timeout", type=float, default=30.0,
                   help="seconds to allow LibreOffice to complete (default: 30)")
    return p


def _report(payload: dict) -> None:
    print(json.dumps(payload, indent=2))


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    path = args.workbook.expanduser().resolve()

    if not path.exists():
        _report({"status": "failed", "path": str(path),
                 "reason": "file does not exist"})
        return 1

    try:
        locate_libreoffice()
    except LibreOfficeNotFound as exc:
        _report({"status": "failed", "path": str(path), "reason": str(exc)})
        return 3

    try:
        bake_workbook(path, args.timeout)
    except subprocess.TimeoutExpired:
        _report({"status": "failed", "path": str(path),
                 "reason": f"LibreOffice exceeded {args.timeout}s timeout"})
        return 4
    except zipfile.BadZipFile as exc:
        _report({"status": "failed", "path": str(path),
                 "reason": f"not a valid xlsx (zip) file: {exc}"})
        return 4
    except BakeFailed as exc:
        _report({"status": "failed", "path": str(path), "reason": str(exc)})
        return 4

    try:
        breakdowns, error_count = collect_error_cells(path)
        formula_count = count_formulas(path)
    except Exception as exc:  # noqa: BLE001
        _report({"status": "failed", "path": str(path),
                 "reason": f"post-bake scan failed: {exc}"})
        return 5

    _report({
        "status": "clean" if error_count == 0 else "errors_present",
        "path": str(path),
        "formula_count": formula_count,
        "error_count": error_count,
        "error_breakdown": [asdict(b) for b in breakdowns],
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
