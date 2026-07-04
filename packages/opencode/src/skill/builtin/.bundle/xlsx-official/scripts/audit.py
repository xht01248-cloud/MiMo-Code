#!/usr/bin/env python3
"""Audit an xlsx: ZIP integrity, XML well-formedness, openpyxl round-trip,
and Excel-error scan.

Prints a JSON audit report and exits 0 on a clean file, 1 if any step
failed, or with a status of ``"warnings"`` (still exit 0) when only the
Excel-error scan turns up hits.

Usage:
    python audit.py workbook.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"ok": False, "reason": "openpyxl is not installed"}),
          file=sys.stderr)
    sys.exit(2)


ERROR_TOKENS: tuple[str, ...] = (
    "#VALUE!", "#DIV/0!", "#REF!", "#NAME?",
    "#NULL!", "#NUM!", "#N/A", "#GETTING_DATA",
)


# --- audit-step dataclasses ----------------------------------------------


@dataclass
class StepResult:
    name: str
    passed: bool
    detail: str = ""
    findings: list[str] = field(default_factory=list)


@dataclass
class AuditReport:
    path: str
    status: str = "ok"                # "ok" | "warnings" | "failed"
    steps: list[StepResult] = field(default_factory=list)

    def append(self, step: StepResult) -> None:
        self.steps.append(step)

    def as_dict(self) -> dict:
        return {
            "path": self.path,
            "status": self.status,
            "steps": [asdict(step) for step in self.steps],
        }


# --- individual probes ----------------------------------------------------


def probe_zip(path: Path) -> tuple[StepResult, list[str]]:
    try:
        with zipfile.ZipFile(path) as archive:
            broken = archive.testzip()
            if broken is not None:
                return (
                    StepResult("zip", False, f"corrupt member: {broken}"),
                    [],
                )
            return (
                StepResult("zip", True, "archive integrity ok"),
                archive.namelist(),
            )
    except zipfile.BadZipFile as exc:
        return StepResult("zip", False, f"not a valid zip: {exc}"), []


def _xml_members(names: Iterable[str]) -> list[str]:
    return [n for n in names if n.endswith(".xml") or n.endswith(".rels")]


def probe_xml(path: Path, members: list[str]) -> StepResult:
    problems: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for name in _xml_members(members):
            try:
                with archive.open(name) as handle:
                    ET.parse(handle)
            except ET.ParseError as exc:
                problems.append(f"{name}: {exc}")
    return StepResult(
        name="xml",
        passed=not problems,
        detail=f"{len(members)} xml/rels parts parsed" if not problems else "xml parse errors",
        findings=problems[:20],
    )


def probe_openpyxl(path: Path) -> StepResult:
    try:
        wb = load_workbook(path, data_only=False)
        try:
            sheet_names = list(wb.sheetnames)
        finally:
            wb.close()
        wb_values = load_workbook(path, data_only=True)
        wb_values.close()
    except Exception as exc:  # noqa: BLE001 — surface any openpyxl error
        return StepResult("openpyxl", False, f"load failed: {exc}")
    return StepResult(
        name="openpyxl",
        passed=True,
        detail=f"loaded in both modes; {len(sheet_names)} worksheet(s)",
    )


def find_error_cells(path: Path) -> StepResult:
    wb = load_workbook(path, data_only=True)
    hits: list[str] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("#"):
                        for token in ERROR_TOKENS:
                            if value == token or value.startswith(token):
                                hits.append(f"{name}!{cell.coordinate}={value}")
                                break
    finally:
        wb.close()

    return StepResult(
        name="excel_errors",
        passed=not hits,
        detail=f"{len(hits)} error cell(s)" if hits else "no error cells",
        findings=hits[:20],
    )


# --- driver ---------------------------------------------------------------


def run_audit(path: Path) -> AuditReport:
    report = AuditReport(path=str(path.resolve()))

    zip_step, members = probe_zip(path)
    report.append(zip_step)
    if not zip_step.passed:
        report.status = "failed"
        return report

    xml_step = probe_xml(path, members)
    report.append(xml_step)
    if not xml_step.passed:
        report.status = "failed"
        return report

    load_step = probe_openpyxl(path)
    report.append(load_step)
    if not load_step.passed:
        report.status = "failed"
        return report

    errors_step = find_error_cells(path)
    report.append(errors_step)
    if not errors_step.passed:
        report.status = "warnings"

    return report


def _parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Audit an xlsx workbook.")
    p.add_argument("workbook", type=Path)
    return p


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if not args.workbook.exists():
        print(json.dumps({"path": str(args.workbook), "status": "failed",
                          "reason": "no such file"}, indent=2))
        return 1

    report = run_audit(args.workbook)
    print(json.dumps(report.as_dict(), indent=2))
    return 0 if report.status in {"ok", "warnings"} else 1


if __name__ == "__main__":
    raise SystemExit(main())
