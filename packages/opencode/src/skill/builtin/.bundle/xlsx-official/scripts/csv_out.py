#!/usr/bin/env python3
"""Flatten an xlsx into one or more delimited text files.

Modes:
    - ``destination`` is a directory  → one file per worksheet
    - ``destination`` is a file path  → single worksheet, named via ``--sheet``

Usage examples:
    python csv_out.py workbook.xlsx dumped/
    python csv_out.py workbook.xlsx summary.csv --sheet Summary
    python csv_out.py workbook.xlsx sheet0.tsv --sheet 0 --delimiter tab

Reads cached values. Run ``bake.py`` first if formulas were freshly written.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is not installed", file=sys.stderr)
    sys.exit(2)


# --- target resolution ----------------------------------------------------


class SheetSelectionError(KeyError):
    """The caller asked for a worksheet that does not exist."""


def resolve_targets(workbook, selector: str | None) -> list[str]:
    """Return the sheet names to export, or raise ``SheetSelectionError``."""
    names = list(workbook.sheetnames)
    if selector is None:
        return names

    if selector.lstrip("-").isdigit():
        index = int(selector)
        if 0 <= index < len(names):
            return [names[index]]
        raise SheetSelectionError(
            f"sheet index {index} out of range 0..{len(names) - 1}"
        )

    if selector in names:
        return [selector]

    raise SheetSelectionError(
        f"worksheet '{selector}' not found; available: {names}"
    )


# --- writer ---------------------------------------------------------------


def _row_is_empty(row: tuple) -> bool:
    return all(cell is None or cell == "" for cell in row)


def emit_rows(worksheet, target: Path, delimiter: str) -> int:
    target.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    with target.open("w", newline="", encoding="utf-8") as fp:
        writer = csv.writer(fp, delimiter=delimiter)
        for row in worksheet.iter_rows(values_only=True):
            if _row_is_empty(row):
                continue
            writer.writerow(["" if cell is None else cell for cell in row])
            rows_written += 1
    return rows_written


def _safe_filename(sheet_name: str) -> str:
    banned = '\\/:*?"<>|'
    cleaned = "".join("_" if ch in banned else ch for ch in sheet_name)
    return cleaned.strip() or "sheet"


def _extension_for(delimiter: str) -> str:
    return "tsv" if delimiter == "\t" else "csv"


# --- CLI ------------------------------------------------------------------


def _parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Flatten an xlsx into CSV/TSV files.")
    p.add_argument("workbook", type=Path)
    p.add_argument("destination", type=Path,
                   help="output file (single sheet) or directory (all sheets)")
    p.add_argument("--sheet",
                   help="worksheet name or 0-based index; omit for all sheets")
    p.add_argument("--delimiter", default=",",
                   help="',' (default), 'tab', or a single character")
    return p


def _resolve_delimiter(raw: str) -> str:
    canon = "\t" if raw.lower() == "tab" else raw
    if len(canon) != 1:
        raise SystemExit("--delimiter must be a single character or 'tab'")
    return canon


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if not args.workbook.exists():
        print(f"no such workbook: {args.workbook}", file=sys.stderr)
        return 1

    delimiter = _resolve_delimiter(args.delimiter)
    workbook = load_workbook(args.workbook, data_only=True, read_only=True)
    try:
        targets = resolve_targets(workbook, args.sheet)

        many_sheets = args.sheet is None or args.destination.is_dir()
        if many_sheets:
            args.destination.mkdir(parents=True, exist_ok=True)
            for sheet_name in targets:
                out = args.destination / f"{_safe_filename(sheet_name)}.{_extension_for(delimiter)}"
                rows = emit_rows(workbook[sheet_name], out, delimiter)
                print(f"{sheet_name} → {out} ({rows} row(s))")
        else:
            [only] = targets
            rows = emit_rows(workbook[only], args.destination, delimiter)
            print(f"{only} → {args.destination} ({rows} row(s))")
    except SheetSelectionError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    finally:
        workbook.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
