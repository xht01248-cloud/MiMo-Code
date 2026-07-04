#!/usr/bin/env python3
"""Emit a JSON snapshot of an xlsx workbook.

Per worksheet the snapshot contains:
    - ``shape``: {"rows": R, "columns": C}
    - ``header_cells``: values from row 1
    - ``column_types``: guessed type per column, based on the sample rows
    - ``formula_count``: number of formula cells
    - ``sample_rows``: up to N stringified data rows (default 5)

Usage:
    python overview.py workbook.xlsx [--rows 5]
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # noqa: F401 - guarded stub for missing dep
    print(
        json.dumps({"ok": False, "reason": "openpyxl is not installed"}),
        file=sys.stderr,
    )
    sys.exit(2)


# --- type inference -------------------------------------------------------


_NUMERIC = {"int", "float"}
_TEMPORAL = {"date", "datetime", "time"}


def infer_column_type(values: list) -> str:
    """Return a coarse type tag ('int', 'float', 'str', 'datetime', ...)."""
    present = [v for v in values if v is not None and v != ""]
    if not present:
        return "empty"
    seen = {type(v).__name__ for v in present}
    if seen == {"bool"}:
        return "bool"
    if seen <= _NUMERIC:
        return "int" if seen == {"int"} else "float"
    if seen & _TEMPORAL:
        return "datetime"
    if seen == {"str"}:
        return "str"
    return "mixed"


# --- snapshot builders ----------------------------------------------------


@dataclass
class WorksheetSnapshot:
    shape: dict
    header_cells: list
    column_types: list[str]
    formula_count: int
    sample_rows: list[list[str]] = field(default_factory=list)


def _count_formulas(ws) -> int:
    total = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                total += 1
    return total


def _read_sample(ws_values, sample_size: int) -> tuple[list, list[list], list[str]]:
    rows_iter = ws_values.iter_rows(values_only=True)
    header = list(next(rows_iter, ()) or ())
    sample: list[list] = []
    for _ in range(sample_size):
        row = next(rows_iter, None)
        if row is None:
            break
        sample.append(list(row))

    column_types: list[str] = []
    for col in range(len(header)):
        column_values = [row[col] if col < len(row) else None for row in sample]
        column_types.append(infer_column_type(column_values))

    stringified = [
        ["" if v is None else str(v) for v in row]
        for row in sample
    ]
    return header, stringified, column_types


def snapshot_worksheet(ws_values, ws_formulas, sample_size: int) -> WorksheetSnapshot:
    header, sample, column_types = _read_sample(ws_values, sample_size)
    return WorksheetSnapshot(
        shape={"rows": ws_values.max_row or 0, "columns": ws_values.max_column or 0},
        header_cells=header,
        column_types=column_types,
        formula_count=_count_formulas(ws_formulas),
        sample_rows=sample,
    )


# --- CLI ------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Emit a JSON snapshot of an xlsx workbook.",
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--rows", type=int, default=5,
                        help="how many data rows to include per worksheet (default: 5)")
    return parser


def _emit(payload: dict) -> None:
    print(json.dumps(payload, indent=2, default=str))


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    if not args.workbook.exists():
        _emit({"ok": False, "reason": f"no such file: {args.workbook}"})
        return 1

    values_wb = load_workbook(args.workbook, data_only=True)
    formula_wb = load_workbook(args.workbook, data_only=False)
    try:
        worksheets = {
            name: asdict(snapshot_worksheet(values_wb[name], formula_wb[name], args.rows))
            for name in values_wb.sheetnames
        }
    finally:
        values_wb.close()
        formula_wb.close()

    _emit({
        "ok": True,
        "path": str(args.workbook.resolve()),
        "worksheet_count": len(worksheets),
        "worksheets": worksheets,
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
